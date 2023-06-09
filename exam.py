import datetime
import os
import time

import openpyxl
import random
from PIL import Image
from openpyxl.utils import get_column_letter

# global variables
apartados = []
duracion = 0


def mostrar_imagen(archivo_imagen):
    try:
        imagen = Image.open(archivo_imagen)
        imagen.show()
    except IOError:
        print("No se pudo cargar la imagen.")


def workbook_unit(unit, cantidad_preguntas, preguntas_seleccionadas, s, hojaN):
    workbook = openpyxl.load_workbook("docs/" + unit)

    # Obtener las hojas del archivo
    hojas = workbook.sheetnames
    # quiere un tema en concreto
    if s == "s":
        hojaN = hojaN.split(",")
        if len(hojaN) > 1:
            nHojas = []
            f = 0
            for i in hojaN:
                nHojas.append(hojas[int(i) - 1])
            hojas = nHojas
        else:
            hojas = [hojas[int(hojaN[0]) - 1]]
    # llamar a la variable global apartados y asignarle hojas
    global apartados
    apartados = hojas
    # sacar una list con max_row de cada hoja
    max_rows = []
    for hoja in hojas:
        sheet = workbook[hoja]
        max_rows.append(sheet.max_row)

    if cantidad_preguntas > min(max_rows):
        cantidad_preguntas = min(max_rows)
    if cantidad_preguntas < 1:
        cantidad_preguntas = 1

    for hoja in hojas:
        num = cantidad_preguntas // len(hojas)
        if hoja == hojas[0]:
            num += cantidad_preguntas % len(hojas)
        # Seleccionar una hoja al azar
        for i in range(num):
            sheet = workbook[hoja]
            # Seleccionar una fila al azar (excluyendo la primera fila de encabezados)
            fila = 1

            # Obtener la pregunta y la respuesta de la fila seleccionada
            pregunta = sheet.cell(row=fila, column=1).value
            respuesta = sheet.cell(row=fila, column=2).value
            img = sheet.cell(row=fila, column=3).value
            it = sheet.cell(row=fila, column=4).value
            year = sheet.cell(row=fila, column=5).value
            justificacion = sheet.cell(row=fila, column=6).value
            if it is None:
                it = 0
            else:
                it = int(it)

            sheet.cell(row=fila, column=4).value = it + 1
            # llevar esta linea al final del documento
            sheet.move_range(f"A1:J1", rows=sheet.max_row)
            sheet.delete_rows(1)
            # Guardar el archivo
            workbook.save("docs/" + unit)

            # Agregar la pregunta seleccionada a la lista
            tema = unit.split(".")[0].replace("u", "")
            preguntas_seleccionadas.append((pregunta, respuesta, hoja, tema, img, year, justificacion))

    workbook.close()


def generar_examen(nombre_archivo, cantidad_preguntas, s, hojaN):
    preguntas_seleccionadas = []
    if cantidad_preguntas > 100:
        cantidad_preguntas = 20
    if nombre_archivo == "r":
        units = ["u7.xlsx", "u8.xlsx", "u9.xlsx", "lab.xlsx"]
        # Cargar el archivo Excel
        # la cantidad de preguntas del lab van a ser siempre 8
        #  el primer parametro el 28% de las preguntas
        #  el segundo parametro el 28% de las preguntas
        #  el tercer parametro el 28% de las preguntas
        #  el cuarto parametro el 16% de las preguntas

        for unit in units:
            if unit == "lab.xlsx":
                num = int(8/60 * cantidad_preguntas)
            else:
                num = int(52/60 * cantidad_preguntas) // (len(units)-1)
            if unit == units[0]:
                num += int(52/60 * cantidad_preguntas) % (len(units)-1)
            workbook_unit(unit, num, preguntas_seleccionadas, s, hojaN)
    else:
        workbook_unit(nombre_archivo, cantidad_preguntas, preguntas_seleccionadas, s, hojaN)

    # Verificar que se hayan seleccionado suficientes preguntas
    # if len(preguntas_seleccionadas) < cantidad_preguntas:
    #     print("Error: No hay suficientes preguntas en el archivo.")
    #     return

    cantidad_preguntas = len(preguntas_seleccionadas)
    # Mezclar las preguntas seleccionadas
    random.shuffle(preguntas_seleccionadas)

    # Imprimir el examen tipo test
    # busar el caracter numero que hay en un string
    tema = nombre_archivo.split(".")[0].replace("u", "")
    print(f"Examen tipo test tema {tema}:")
    print("-----------------")

    aciertos = 0
    fallos = 0

    for i in range(cantidad_preguntas):
        pregunta = preguntas_seleccionadas[i][0]
        respuesta = preguntas_seleccionadas[i][1]
        hoja = preguntas_seleccionadas[i][2]
        tema = preguntas_seleccionadas[i][3]
        img = preguntas_seleccionadas[i][4]
        year = preguntas_seleccionadas[i][5]
        justificacion = preguntas_seleccionadas[i][6]

        # de pregunta quitar todo lo que haya antes del primer punto, puede haber mas de un punto
        pregunta = pregunta.split(".")
        # quitar el primer elemento de la lista
        pregunta[0].strip()
        if pregunta[0].isdigit():
            pregunta.pop(0)
        # volver a unir la lista en un string
        pregunta = ".".join(pregunta)
        pregunta = pregunta.strip()

        # poner preguntas de color azul
        # print verde
        # linea que separa
        print("-------------------------------------------")
        # color rojo para los fallos
        print(f"\033[0mAciertos: {aciertos}/{cantidad_preguntas}  |  Fallos: {fallos}/{cantidad_preguntas}")

        if year is not None:
            year = f"[Año: {year}]"
        else:
            year = ""
        # color azul para las preguntas
        print("\033[94m")
        print(f"Pregunta([Unidad {tema}]{year} - {hoja}) [{i + 1} de {cantidad_preguntas}]:\033[0m")
        # hacer bold la pregunta
        # color amarillo para las preguntas
        print("\033[93m")
        print(pregunta)
        print("\033[0m")
        t = False
        imagen = None
        # hacer selector de verdadero o falso con teclado
        # 
        if img is not None and img != "" and img.startswith("img"):
            img = "img/" + img
            imagen = Image.open(img)
            # abrir en segundo plano
            imagen.show()
            respuesta_usuario = input("Ingrese su respuesta (V o F): ")
            os.system("taskkill /f /im PhotosApp.exe >nul 2>&1")
        else:
            respuesta_usuario = input("Ingrese su respuesta (V o F): ")

        if respuesta_usuario.upper() == respuesta.upper():
            # color verde para las respuestas correctas
            print("\033[92m")
            print("¡Respuesta correcta!")
            print("\033[0m")
            aciertos += 1
        # en el caso de que no responda, es decir haga un enter sin escribir nada ni sumar ni restar
        elif respuesta_usuario == "":
            print("No has respondido.")
        else:
            fallos += 1
            # color rojo para las respuestas incorrectas
            print("\033[91m")
            print("Respuesta incorrecta.")
            print("\033[0m")

        # si he acertado, no muestro la respuesta correcta
        if respuesta_usuario.upper() == respuesta.upper():
            if justificacion is not None and justificacion != "":
                print(f"\033[94mJustificación: {justificacion}\033[0m")
            continue
        # Mostrar la respuesta correcta
        print(f"Respuesta correcta: {respuesta}")
        if justificacion is not None and justificacion != "":
            print(f"\033[94mJustificación: {justificacion}\033[0m")

        print()

    # Cerrar el archivo

    # Mostrar resultados del examen
    print("Resultados del examen:")
    print("---------------------")
    # sabiendo que una respuesta correcta vale 1 punto y una incorrecta resta 1 punto
    nota = ((aciertos - fallos) / cantidad_preguntas) * 10
    # color verde para las notas mayores o iguales a 5
    if nota >= 5:
        print("\033[92m")
    else:
        print("\033[91m")
    print(f"Preguntas respondidas correctamente: {aciertos}/{cantidad_preguntas}: {nota:.2f}")
    print(f"Las preguntas respondidas incorrectamente restan 1 punto");

    #     guardar los resultados en un excel
    #     primera columna: tema
    #     segunda columna: fecha
    #     tercera columna: numero de aciertos
    #     cuarta columna: numero de fallos
    #     quinta columna: nota
    fecha = datetime.datetime.now()
    fecha = fecha.strftime("%d/%m/%Y %H:%M:%S")
    if nombre_archivo == "u7.xlsx":
        tema = "7"
    elif nombre_archivo == "u8.xlsx":
        tema = "8"
    elif nombre_archivo == "u9.xlsx":
        tema = "9"
    elif nombre_archivo == "lab.xlsx":
        tema = "lab"
    else:
        tema = "r"
    # abrir el archivo excel resultados.xlsx si no existe lo crea
    try:
        workbook = openpyxl.load_workbook("resultados.xlsx")
    except:
        workbook = openpyxl.Workbook()

    # seleccionar la hoja, si no existe la crea
    try:
        sheet = workbook["resultados"]
    except:
        sheet = workbook.create_sheet("resultados")
    #     añadir en la primera fila los encabezados
    # si hay otras hojas que no sea resultados, eliminarlas

    # obtener el numero de la ultima fila
    max_filas = sheet.max_row
    # escribir en la ultima fila
    # cast tema into a int
    if tema == "r":
        tema = 789
    elif tema == "lab":
        tema = 45
    else:
        tema = int(tema)

    # set the list apartados to a string
    global apartados
    apartados = ", ".join(apartados)
    global duracion
    duracion = time.time() - duracion
    duracion = round(duracion, 2)
    if max_filas == 1:
        sheet.cell(row=max_filas, column=1).value = "Tema"
        sheet.cell(row=max_filas, column=2).value = "Apartado"
        sheet.cell(row=max_filas, column=3).value = "Duración"
        sheet.cell(row=max_filas, column=4).value = "Fecha"
        sheet.cell(row=max_filas, column=5).value = "Preguntas"
        sheet.cell(row=max_filas, column=6).value = "Aciertos"
        sheet.cell(row=max_filas, column=7).value = "Fallos"
        sheet.cell(row=max_filas, column=8).value = "Nota"
    sheet.cell(row=max_filas + 1, column=1).value = tema
    sheet.cell(row=max_filas + 1, column=2).value = apartados
    sheet.cell(row=max_filas + 1, column=3).value = duracion
    sheet.cell(row=max_filas + 1, column=4).value = fecha
    sheet.cell(row=max_filas + 1, column=5).value = cantidad_preguntas
    sheet.cell(row=max_filas + 1, column=6).value = aciertos
    sheet.cell(row=max_filas + 1, column=7).value = fallos
    sheet.cell(row=max_filas + 1, column=8).value = nota
    # guardar el archivo
    workbook.save("resultados.xlsx")
    workbook.close()
    print("\033[0m")
    print("Resultados guardados en resultados.xlsx")
    print("\033[0m")

    # poner todas las preguntas que ha fallado aqui:

    #     preguntar si quiere hacer otro examen
    #     si quiere hacer otro examen, volver al principio
    #     si no quiere hacer otro examen, salir del programa
    if input("¿Quieres hacer otro examen? (s/n): ") == "s":
        # limpiar la pantalla
        os.system("cls")
        main()
    else:
        input("Pulsa para salir")


def main():
    # Ejemplo de uso
    # poner en azul un titulo
    print("\033[94m")
    print("Generador de exámenes tipo test CSD parcial 2")
    print("\033[0m")

    tema = input("Ingrese el tema del examen (7,8,9) o practs ([l]lab) y si es de repaso (r): ")
    if tema == "7":
        archivo_excel = "u7.xlsx"
    elif tema == "8":
        archivo_excel = "u8.xlsx"
    elif tema == "9":
        archivo_excel = "u9.xlsx"
    elif tema == "r":
        archivo_excel = "r"
    elif tema == "l":
        archivo_excel = "lab.xlsx"
    else:
        archivo_excel = "u7.xlsx"

    s = "n"
    if archivo_excel != "r":
        s = input("¿Quieres un contenido en concreto? (s/n): ")

    if s == "s":
        libro = openpyxl.load_workbook("docs/" + archivo_excel)
        hojas = libro.sheetnames
        i = 1
        numero_hojas = []
        for hoja in hojas:
            print(f"{i}) {hoja}: {libro[hoja].max_row} preguntas")
            numero_hojas.append(libro[hoja].max_row)
            i += 1
        hojaN = input("¿Qué contenido quieres?(1,2,..): ")
        cantidad_preguntas_deseada = input("Ingrese la cantidad de preguntas que desea: ")
        if cantidad_preguntas_deseada == "":
            cantidad_preguntas_deseada = numero_hojas[int(hojaN) - 1]
        else:
            cantidad_preguntas_deseada = int(cantidad_preguntas_deseada)
    else:
        cantidad_preguntas_deseada = int(input("Ingrese la cantidad de preguntas que desea: "))
        hojaN = 0

    # empezar el timer de la duracion del examen get the current time
    global duracion
    duracion = time.time()
    generar_examen(archivo_excel, cantidad_preguntas_deseada, s, hojaN)


main()
